import openpyxl 

# 1) 엑셀만들기
wb = openpyxl.Workbook()

# 2) 엑셀 워크시트 만들기
ws = wb.create_sheet('오징어게임')

# 3) 데이터 추가하기
#    A  B  C  D
# 1 A1 B1
# 2
# 3
# 4
ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = 1
ws['B2'] = '오일남'

# 4) 엑셀 저장하기
wb.save(r'C:\Users\Administrator\Documents\GitHub\python_Crolling\02_파이썬엑셀다루기\참가자_data.xlsx')